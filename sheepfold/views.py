from datetime import timedelta

from django.db import models
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
import icalendar

from .models import Sheep, BirthEvent, Milk, HealthRecord, CalendarEvent, MilkAnalysis
from .services import assign_groups
from .serializers import (
    MilkSerializer, SheepData, BirthEventSerializer, HealthRecordSerializer,
    CalendarEventSerializer, MilkAnalysisSerializer,
)
from accounts.helpers import (
    get_active_farm, get_user_role, role_required, api_role_required,
)
from accounts.models import Farm, ROLE_FARMER, ROLE_ANALYST, ROLE_MANAGER


def _require_farm(request):
    """Return the active farm or None."""
    return get_active_farm(request)


# ---------------------------------------------------------------------------
# Page views
# ---------------------------------------------------------------------------

@login_required(login_url='login')
@role_required(ROLE_FARMER, ROLE_MANAGER)
def homepage(request):
    farm = _require_farm(request)
    if farm is None:
        return redirect('select_farm')

    sheep = Sheep.objects.filter(farm=farm)
    today = timezone.localdate()

    # Recent activity: last 10 items across milk, birth events
    # Sorted by date (then by id as tiebreaker within same date)
    recent_milk = (
        Milk.objects.filter(sheep__farm=farm)
        .select_related('sheep')
        .order_by('-date', '-id')[:10]
    )
    recent_births = (
        BirthEvent.objects.filter(mother__farm=farm)
        .select_related('mother')
        .prefetch_related('lambs')
        .order_by('-date', '-id')[:10]
    )

    activity = []
    for m in recent_milk:
        activity.append({
            'icon': 'bi-droplet-fill',
            'color': '#3b82f6',
            'text': f'Milk recorded for #{m.sheep.earing} — {m.milk}L on {m.date}',
            'date': m.date,
            'id': m.id,
        })
    for b in recent_births:
        lamb_count = b.lambs.count()
        activity.append({
            'icon': 'bi-lightning-fill',
            'color': '#f59e0b',
            'text': f'Birth event: #{b.mother.earing} — {lamb_count} lamb(s) on {b.date}',
            'date': b.date,
            'id': b.id,
        })
    activity.sort(key=lambda x: (x['date'], x['id']), reverse=True)
    activity = activity[:10]

    # Upcoming calendar events (next 5)
    upcoming_events = (
        CalendarEvent.objects.filter(farm=farm, start__gte=today)
        .order_by('start')[:5]
    )

    return render(request, "homepage.html", {
        "sheep": sheep,
        "farm": farm,
        "activity": activity,
        "upcoming_events": upcoming_events,
        "milk_count": Milk.objects.filter(sheep__farm=farm).count(),
        "birth_count": BirthEvent.objects.filter(mother__farm=farm).count(),
        "health_count": HealthRecord.objects.filter(farm=farm).count(),
    })


@login_required(login_url='login')
@role_required(ROLE_FARMER, ROLE_MANAGER)
def sheep_create(request):
    farm = _require_farm(request)
    if farm is None:
        return redirect('select_farm')
    return render(request, "sheep_form.html")


@login_required(login_url='login')
@role_required(ROLE_FARMER, ROLE_MANAGER)
def milking(request):
    farm = _require_farm(request)
    if farm is None:
        return redirect('select_farm')
    milk = Milk.objects.filter(sheep__farm=farm)
    sheep = Sheep.objects.filter(farm=farm)
    return render(request, "milking.html", {"sheep": sheep, "milk": milk})


@login_required(login_url='login')
@role_required(ROLE_FARMER, ROLE_MANAGER)
def lamping(request):
    farm = _require_farm(request)
    if farm is None:
        return redirect('select_farm')

    return render(request, 'lamping.html')


@login_required(login_url='login')
@role_required(ROLE_FARMER, ROLE_MANAGER)
def health_view(request):
    farm = _require_farm(request)
    if farm is None:
        return redirect('select_farm')
    sheep = Sheep.objects.filter(farm=farm)
    return render(request, "health.html", {"sheep": sheep})


@login_required(login_url='login')
@role_required(ROLE_FARMER, ROLE_MANAGER)
def bulk_milking(request):
    farm = _require_farm(request)
    if farm is None:
        return redirect('select_farm')
    return render(request, "bulk_milking.html")


@login_required(login_url='login')
@role_required(ROLE_FARMER, ROLE_ANALYST, ROLE_MANAGER)
def milk_analysis_view(request):
    farm = _require_farm(request)
    if farm is None:
        return redirect('select_farm')
    role = get_user_role(request.user)
    can_create = role in (ROLE_ANALYST, ROLE_MANAGER) or request.user.is_superuser
    can_edit = can_create
    can_delete = role == ROLE_MANAGER or request.user.is_superuser
    return render(request, "milk_analysis.html", {
        "can_create": can_create,
        "can_edit": can_edit,
        "can_delete": can_delete,
    })


@login_required(login_url='login')
@role_required(ROLE_FARMER, ROLE_MANAGER)
def calendar_view(request):
    farm = _require_farm(request)
    if farm is None:
        return redirect('select_farm')
    return render(request, "calendar.html", {"calendar_token": farm.calendar_token})


# ---------------------------------------------------------------------------
# API views
# ---------------------------------------------------------------------------

@login_required(login_url='login')
@api_view(['GET', 'POST'])
@api_role_required(ROLE_FARMER, ROLE_MANAGER)
def milking_api(request):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        milk = Milk.objects.filter(sheep__farm=farm)
        serializer = MilkSerializer(milk, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = MilkSerializer(data=request.data, context={'farm': farm})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@login_required(login_url='login')
@api_view(['PUT', 'DELETE'])
@api_role_required(ROLE_FARMER, ROLE_MANAGER)
def milk_detail_api(request, pk):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    milk = get_object_or_404(Milk, pk=pk, sheep__farm=farm)

    if request.method == 'DELETE':
        milk.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    serializer = MilkSerializer(milk, data=request.data, partial=True, context={'farm': farm})
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@login_required(login_url='login')
@api_view(['PUT', 'DELETE'])
@api_role_required(ROLE_FARMER, ROLE_MANAGER)
def sheep_detail_api(request, pk):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    sheep = get_object_or_404(Sheep, pk=pk, farm=farm)

    if request.method == 'DELETE':
        sheep.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    serializer = SheepData(sheep, data=request.data, partial=True, context={'farm': farm})
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@login_required(login_url='login')
@api_view(['GET', 'POST'])
@api_role_required(ROLE_FARMER, ROLE_MANAGER)
def birthevent_api(request):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        events = (
            BirthEvent.objects.filter(mother__farm=farm)
            .select_related('mother')
            .prefetch_related('lambs')
            .order_by('-date')
        )
        serializer = BirthEventSerializer(events, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = BirthEventSerializer(data=request.data, context={'farm': farm})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@login_required(login_url='login')
@api_view(['PUT', 'DELETE'])
@api_role_required(ROLE_FARMER, ROLE_MANAGER)
def birthevent_detail_api(request, pk):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    event = get_object_or_404(BirthEvent, pk=pk, mother__farm=farm)

    if request.method == 'DELETE':
        event.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    # Only allow editing date and notes (mother/lambs are complex)
    data = {}
    if 'date' in request.data:
        data['date'] = request.data['date']
    if 'notes' in request.data:
        data['notes'] = request.data['notes']

    for key, value in data.items():
        setattr(event, key, value)
    event.save()

    serializer = BirthEventSerializer(event)
    return Response(serializer.data)


@login_required(login_url='login')
@api_view(['GET', 'POST'])
@api_role_required(ROLE_FARMER, ROLE_MANAGER)
def health_api(request):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        records = HealthRecord.objects.filter(farm=farm).select_related('sheep').order_by('-date')
        serializer = HealthRecordSerializer(records, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = HealthRecordSerializer(data=request.data, context={'farm': farm})
        if serializer.is_valid():
            serializer.save(farm=farm)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@login_required(login_url='login')
@api_view(['PUT', 'DELETE'])
@api_role_required(ROLE_FARMER, ROLE_MANAGER)
def health_detail_api(request, pk):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    record = get_object_or_404(HealthRecord, pk=pk, farm=farm)

    if request.method == 'DELETE':
        record.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    serializer = HealthRecordSerializer(record, data=request.data, partial=True, context={'farm': farm})
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@login_required(login_url='login')
@api_view(['GET'])
@api_role_required(ROLE_FARMER, ROLE_MANAGER)
def sheep_profile_api(request, pk):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    sheep = get_object_or_404(Sheep, pk=pk, farm=farm)

    # Basic info
    info = {
        "id": sheep.id,
        "earing": sheep.earing,
        "gender": sheep.get_gender_display(),
        "gender_code": sheep.gender,
        "birthdate": sheep.birthdate,
        "is_active": sheep.is_active,
        "mother": sheep.mother.earing if sheep.mother else None,
        "mother_id": sheep.mother_id,
        "father": sheep.father.earing if sheep.father else None,
        "father_id": sheep.father_id,
        "group": sheep.group,
        "ready_for_birth": sheep.ready_for_birth,
    }

    # Children (as mother or father)
    children = list(
        Sheep.objects.filter(
            models.Q(mother=sheep) | models.Q(father=sheep)
        ).values('id', 'earing', 'gender', 'birthdate').distinct()
    )

    # Health records (individual + batch)
    health = list(
        HealthRecord.objects.filter(
            farm=farm
        ).filter(
            models.Q(sheep=sheep) | models.Q(is_batch=True)
        ).order_by('-date').values(
            'id', 'date', 'record_type', 'title', 'notes', 'next_due', 'is_batch'
        )[:50]
    )

    # Milk records
    milk = list(
        Milk.objects.filter(sheep=sheep).order_by('-date').values(
            'id', 'date', 'milk', 'is_active'
        )[:50]
    )

    # Birth events (as mother)
    births_as_mother = []
    for be in BirthEvent.objects.filter(mother=sheep).prefetch_related('lambs').order_by('-date')[:20]:
        births_as_mother.append({
            'id': be.id,
            'date': be.date,
            'notes': be.notes,
            'lambs': list(be.lambs.values('id', 'earing', 'gender')),
        })

    # Birth event where this sheep was born (as a lamb)
    born_in = None
    birth_event = sheep.birth_event.select_related('mother').first()
    if birth_event:
        born_in = {
            'id': birth_event.id,
            'date': birth_event.date,
            'mother': birth_event.mother.earing,
        }

    return Response({
        "info": info,
        "children": children,
        "health": health,
        "milk": milk,
        "births_as_mother": births_as_mother,
        "born_in": born_in,
    })


@login_required(login_url='login')
@role_required(ROLE_FARMER, ROLE_MANAGER)
def sheep_export_excel(request, pk):
    farm = get_active_farm(request)
    if farm is None:
        return redirect('select_farm')

    sheep = get_object_or_404(Sheep, pk=pk, farm=farm)

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='2D6A4F')
    center = Alignment(horizontal='center')

    def style_header_row(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # --- Sheet 1: Info ---
    ws_info = wb.active
    ws_info.title = 'Info'
    ws_info.append(['Field', 'Value'])
    style_header_row(ws_info)
    rows = [
        ('Earing', sheep.earing),
        ('Gender', sheep.get_gender_display()),
        ('Birthdate', str(sheep.birthdate) if sheep.birthdate else ''),
        ('Active', 'Yes' if sheep.is_active else 'No'),
        ('Group', sheep.get_group_display() if sheep.group else ''),
        ('Ready for Birth', 'Yes' if sheep.ready_for_birth else 'No'),
        ('Mother', sheep.mother.earing if sheep.mother else ''),
        ('Father', sheep.father.earing if sheep.father else ''),
        ('Farm', farm.name),
    ]
    for row in rows:
        ws_info.append(row)
    auto_width(ws_info)

    # --- Sheet 2: Milk ---
    ws_milk = wb.create_sheet('Milk Records')
    ws_milk.append(['Date', 'Liters', 'Active'])
    style_header_row(ws_milk)
    for m in Milk.objects.filter(sheep=sheep).order_by('-date'):
        ws_milk.append([str(m.date), float(m.milk), 'Yes' if m.is_active else 'No'])
    auto_width(ws_milk)

    # --- Sheet 3: Health ---
    ws_health = wb.create_sheet('Health Records')
    ws_health.append(['Date', 'Type', 'Title', 'Notes', 'Next Due', 'Batch'])
    style_header_row(ws_health)
    for h in HealthRecord.objects.filter(farm=farm).filter(
        models.Q(sheep=sheep) | models.Q(is_batch=True)
    ).order_by('-date'):
        ws_health.append([
            str(h.date), h.record_type, h.title, h.notes,
            str(h.next_due) if h.next_due else '', 'Yes' if h.is_batch else 'No',
        ])
    auto_width(ws_health)

    # --- Sheet 4: Births (as mother) ---
    ws_births = wb.create_sheet('Births as Mother')
    ws_births.append(['Date', 'Lambs', 'Notes'])
    style_header_row(ws_births)
    for be in BirthEvent.objects.filter(mother=sheep).prefetch_related('lambs').order_by('-date'):
        lambs = ', '.join(l.earing for l in be.lambs.all())
        ws_births.append([str(be.date), lambs, be.notes])
    auto_width(ws_births)

    # --- Response ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sheep_{sheep.earing}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login')
@api_view(['GET', 'POST'])
@api_role_required(ROLE_FARMER, ROLE_MANAGER)
def sheep_data_api(request):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        sheep = Sheep.objects.filter(farm=farm)
        serializer = SheepData(sheep, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = SheepData(data=request.data, context={'farm': farm})
        if serializer.is_valid():
            serializer.save(farm=farm)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@login_required(login_url='login')
@api_view(['GET', 'POST'])
@api_role_required(ROLE_FARMER, ROLE_MANAGER)
def calendar_data_api(request):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        all_farms = request.query_params.get('all_farms') == 'true'
        role = get_user_role(request.user)
        if all_farms and (role == ROLE_MANAGER or request.user.is_superuser):
            events = CalendarEvent.objects.select_related('farm').all()
        else:
            events = CalendarEvent.objects.select_related('farm').filter(farm=farm)
        serializer = CalendarEventSerializer(events, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        serializer = CalendarEventSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(farm=farm)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@login_required(login_url='login')
@api_view(['PUT', 'DELETE'])
@api_role_required(ROLE_FARMER, ROLE_MANAGER)
def calendar_detail_api(request, pk):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    event = get_object_or_404(CalendarEvent, pk=pk, farm=farm)

    if request.method == 'DELETE':
        event.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    serializer = CalendarEventSerializer(event, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ---------------------------------------------------------------------------
# Milk Analysis API
# ---------------------------------------------------------------------------

@login_required(login_url='login')
@api_view(['GET', 'POST'])
@api_role_required(ROLE_FARMER, ROLE_ANALYST, ROLE_MANAGER)
def milk_analysis_api(request):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        analyses = MilkAnalysis.objects.filter(farm=farm).select_related('created_by')
        serializer = MilkAnalysisSerializer(analyses, many=True, context={'request': request})
        return Response(serializer.data)

    # POST — analyst or manager only
    role = get_user_role(request.user)
    if not (request.user.is_superuser or role in (ROLE_ANALYST, ROLE_MANAGER)):
        return Response(
            {"detail": "You don't have permission to upload analyses."},
            status=status.HTTP_403_FORBIDDEN,
        )

    serializer = MilkAnalysisSerializer(data=request.data, context={'request': request})
    if serializer.is_valid():
        serializer.save(farm=farm, created_by=request.user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@login_required(login_url='login')
@api_view(['GET', 'PUT', 'DELETE'])
@api_role_required(ROLE_FARMER, ROLE_ANALYST, ROLE_MANAGER)
def milk_analysis_detail_api(request, pk):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    analysis = get_object_or_404(MilkAnalysis, pk=pk, farm=farm)
    role = get_user_role(request.user)
    is_super = request.user.is_superuser

    if request.method == 'GET':
        serializer = MilkAnalysisSerializer(analysis, context={'request': request})
        return Response(serializer.data)

    if request.method == 'DELETE':
        if not (is_super or role == ROLE_MANAGER):
            return Response(
                {"detail": "Only managers can delete analyses."},
                status=status.HTTP_403_FORBIDDEN,
            )
        analysis.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    # PUT — analyst or manager only
    if not (is_super or role in (ROLE_ANALYST, ROLE_MANAGER)):
        return Response(
            {"detail": "You don't have permission to edit analyses."},
            status=status.HTTP_403_FORBIDDEN,
        )

    serializer = MilkAnalysisSerializer(
        analysis, data=request.data, partial=True, context={'request': request}
    )
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ---------------------------------------------------------------------------
# Genealogy API
# ---------------------------------------------------------------------------

@login_required(login_url='login')
@api_view(['GET'])
@api_role_required(ROLE_FARMER, ROLE_MANAGER)
def genealogy_api(request):
    farm = get_active_farm(request)
    if farm is None:
        return Response({"detail": "No active farm."}, status=status.HTTP_403_FORBIDDEN)

    sheep_qs = Sheep.objects.filter(farm=farm).select_related('mother', 'father')
    nodes = []
    for s in sheep_qs:
        nodes.append({
            'id': s.id,
            'earing': s.earing,
            'gender': s.gender,
            'birthdate': str(s.birthdate) if s.birthdate else None,
            'is_active': s.is_active,
            'mother_id': s.mother_id,
            'father_id': s.father_id,
            'group': s.group,
        })
    return Response(nodes)


@login_required(login_url='login')
@role_required(ROLE_FARMER, ROLE_MANAGER)
def genealogy_view(request):
    farm = _require_farm(request)
    if farm is None:
        return redirect('select_farm')
    return render(request, "genealogy.html")


# ---------------------------------------------------------------------------
# Group recalculation
# ---------------------------------------------------------------------------

@login_required(login_url='login')
@api_view(['POST'])
@api_role_required(ROLE_FARMER, ROLE_MANAGER)
def recalculate_groups_api(request):
    farm = get_active_farm(request)
    if farm is None:
        return Response({'error': 'No active farm'}, status=status.HTTP_400_BAD_REQUEST)
    assign_groups(farm)
    return Response({'status': 'ok'})


# ---------------------------------------------------------------------------
# iCal feed (no login — authenticated by secret token)
# ---------------------------------------------------------------------------

def calendar_feed(request, token):
    farm = get_object_or_404(Farm, calendar_token=token)

    cal = icalendar.Calendar()
    cal.add('prodid', '-//Sheeper//Farm Calendar//EN')
    cal.add('version', '2.0')
    cal.add('x-wr-calname', f'Sheeper - {farm.name}')

    for event in CalendarEvent.objects.filter(farm=farm):
        vevent = icalendar.Event()
        vevent.add('uid', f'sheeper-event-{event.pk}@sheeper')
        vevent.add('summary', event.title)
        vevent.add('dtstart', event.start)
        if event.end:
            vevent.add('dtend', event.end)
        else:
            vevent.add('dtend', event.start + timedelta(days=1))
        cal.add_component(vevent)

    response = HttpResponse(cal.to_ical(), content_type='text/calendar; charset=utf-8')
    return response
